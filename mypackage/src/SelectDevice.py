import os
import re
import time
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from PyQt5 import QtCore
from PyQt5.QtCore import QDir, ws
from PyQt5.QtWidgets import QWidget, QPushButton, QLineEdit, QTextEdit, QVBoxLayout, QHBoxLayout, QFileDialog, \
    QComboBox, QLabel
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

class SelectDevice(QWidget):
    def __init__(self):
        super(SelectDevice, self).__init__()
        self.error = '<font color="red">{}</font>'
        self.warning = '<font color="orange">{}</font>'
        self.valid = '<font color="green">{}</font>'
        self.normal = '<font color="black">{}</font>'
        #设置文件
        self.set_file_path = None

        #刷选结果
        self.df = pd.DataFrame()

        #实例化按键
        self.btn = QPushButton("打开需要刷选的目录")
        self.btn.setStyleSheet("background-color: rgb(225,225,225); color: black;")
        self.btn.setFixedHeight(40)
        self.btn.clicked.connect(self.get_dir)

        #文本框设定
        self.current_path = QtCore.QDir.currentPath()
        self.line_edit_path = QLineEdit(str(self.current_path))
        self.line_edit_path.setFixedHeight(40)
        self.line_edit_path.setStyleSheet("QLineEdit { background-color: white; }")

        #加载所有机型
        self.label_device = QLabel("选定机型")
        self.label_device.setStyleSheet("background-color: rgb(225,225,225); color: black;")
        self.label_device.setFixedHeight(40)
        self.cb = QComboBox()
        self.cb.setFixedHeight(40)
        self.cb.setStyleSheet("QComboBox { background-color: white; }")

        #开始按键
        self.push_button_start = QPushButton('开始')
        self.push_button_start.setStyleSheet("background-color: rgb(255,255,255); color: black;")
        self.push_button_start.setFixedSize(100, 40)
        self.push_button_start.clicked.connect(self.start_button)

        #实例化textEdit并加入布局
        self.text_edit = QTextEdit()
        self.text_edit.setStyleSheet("QTextEdit { background-color: white; }")

        # 读设置信息
        set_path = self.current_path + '/setting/'
        if not os.path.isdir(set_path):
            os.mkdir(set_path)
            print("set文件不存在")
            self.text_edit.append(self.error.format("请先在setting目录下添加设置文件❌"))
        else:
            self.set_file_path = Path(set_path + 'setting.xlsx')
            if self.set_file_path.is_file():
                try:
                    excel_file = pd.ExcelFile(self.set_file_path)
                    sheet_names = excel_file.sheet_names
                    self.cb.addItems(sheet_names)

                    print('set文件存在')
                    self.text_edit.append(self.valid.format("set文件存在✅"))
                except FileNotFoundError as e:
                    print(f"设置文件打开失败: {e}")
                    self.text_edit.append(self.error.format(f"设置文件打开失败: {e}❌"))
            else:
                print("set文件不存在")
                self.text_edit.append(self.error.format("设置文件不存在❌"))

        #布局
        stack_main_layout = QVBoxLayout()
        stack_layout_1 = QHBoxLayout()
        stack_layout_2 = QHBoxLayout()

        #并加入布局
        stack_layout_1.addWidget(self.btn)
        stack_layout_1.addWidget(self.line_edit_path)
        stack_layout_1.addWidget(self.label_device)
        stack_layout_1.addWidget(self.cb)
        stack_layout_1.addWidget(self.push_button_start)
        stack_layout_2.addWidget(self.text_edit)
        stack_main_layout.addLayout(stack_layout_1)
        stack_main_layout.addLayout(stack_layout_2)
        self.setLayout(stack_main_layout)


    def get_dir(self):
        # # 1. 实例化对话框
        # dialog = QFileDialog()
        # # 2. 设置标题
        # dialog.setWindowTitle("选择包含log文件的目录")
        # # 3. 关键设置：设置为目录模式
        # dialog.setFileMode(QFileDialog.Directory)
        # # 4. 设置过滤器
        # dialog.setNameFilters(["log文件 (*.log)"])
        # # # 5. (可选) 强制使用 Qt 原生样式而非系统样式，以确保过滤器在所有系统上表现一致
        # dialog.setOption(QFileDialog.DontUseNativeDialog, True)
        #
        # # 6. 执行对话框
        # if dialog.exec_():
        #     # 获取选中的文件夹路径
        #     selected_dir = dialog.selectedFiles()[0]
        #     print(f"你选择的文件夹是: {selected_dir}")
        #
        #     # 你可以在这里进一步确认该文件夹下是否有你想要的文件
        #     # 例如检查该目录下是否存在 .xlsx 文件
        #     dir_obj = QDir(selected_dir)
        #     files = dir_obj.entryList(["*.log"])
        #     if files:
        #         print(f"该目录下包含 {len(files)} 个 log 文件")
        #         self.current_path = selected_dir
        #         self.line_edit_path.setText(selected_dir)
        #     else:
        #         print("该目录下没有找到 log 文件")

        dialog = QFileDialog()
        dialog.options = QFileDialog.Options()
        dialog.options |= QFileDialog.ShowDirsOnly
        # dialog.setFileMode(QFileDialog.ExistingFiles)
        folder_path = dialog.getExistingDirectory(self, "选择文件夹", options=dialog.options)

        if folder_path:
            print(f"选择的文件夹：{folder_path}")
            self.current_path = folder_path
            self.line_edit_path.setText(folder_path)

    def start_button(self):
        print("开始")
        temp_path = QtCore.QDir(self.line_edit_path.text())
        path = temp_path.absolutePath()
        print('目录:'+path)
        self.text_edit.clear()
        # self.text_edit.append('目录:'+path)
        if not os.path.isdir(path):
            print("错误")
            self.text_edit.append(self.error.format("目录错误❌"))
        else:
            print("正常")
            self.text_edit.append(self.valid.format("目录正常✅"))
            txt_files = [file for file in os.listdir(path) if file.endswith(".log")]
            if len(txt_files) > 0:
                self.text_edit.append(self.valid.format("目录存在log文件✅"))
                self.log_process(path, txt_files)
            else:
                self.text_edit.append(self.error.format("未找到log文件❌"))

    @staticmethod
    def get_timestamp(line):
        timestamp = 0
        match = re.search(r'\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}\.\d{3}', line)
        if match:
            time_str = match.group()
            # 1. 将字符串解析为 datetime 对象
            # %Y-%m-%d %H:%M:%S.%f 分别对应 年-月-日 时:分:秒.毫秒
            dt_obj = datetime.strptime(time_str, "%Y-%m-%d %H:%M:%S.%f")

            # 2. 将 datetime 对象转换为时间戳
            timestamp = dt_obj.timestamp()
            print(timestamp)

        return timestamp


    def log_process(self, path, log_files):
        sheet_name = self.cb.currentText()
        print(sheet_name)
        try:
            set_data = pd.read_excel(self.set_file_path, sheet_name=sheet_name, dtype=str)  # 以字符串形式打开并读取excel表格
            pd_set_data = pd.DataFrame(set_data)
            print(pd_set_data)

            self.text_edit.append(self.valid.format(sheet_name+"配置文件打开✅"))
        except FileNotFoundError as e:
            self.text_edit.append(self.error.format(sheet_name+"配置文件打开❌"))
            return

        pd_data = pd.DataFrame()
        #通过修改index属性更改行名称
        pd_data.index = ['通用-A通道校机差值', '通用-B通道校机差值', '通用-B通道初始增量',
            'UL烟雾-灵敏度', 'UL烟雾-增量比值', 'UL烟雾-比值变化率', 'UL烟雾-计数', 'UL烟雾-PPM', 'UL烟雾-平均值', 'UL烟雾-方差',
            'PU烟雾-灵敏度', 'PU烟雾-增量比值', 'PU烟雾-比值变化率', 'PU烟雾-计数', 'PU烟雾-PPM', 'PU烟雾-平均值', 'PU烟雾-方差',
            '油烟-增量比值', '油烟-比值变化率', '油烟-计数', '油烟-PPM', '油烟-平均值',  '油烟-方差', '油烟-结束时红光增量',
            '混合烟-数组最小值', '混合烟-数组最大值', '混合烟-增量比值', '混合烟-比值变化率']

        #设备状态
        states_index = 0
        states_bit_start = 12 + states_index * 3
        # A增量
        increment_a_index = 17
        increment_a_bit_start = 12 + increment_a_index * 3
        # B增量
        increment_b_index = 34
        increment_b_bit_start = 12 + increment_b_index * 3
        # 增量比值
        increment_ration_list = []
        increment_index = 36
        increment_bit_start = 12 + increment_index*3
        # 比值变化率
        ration_of_change_list = []
        ration_index = 38
        ration_bit_start = 12 + ration_index*3
        # 计数
        rise_cnt_list = []
        rise_index = 40
        rise_bit_start = 12 + rise_index*3
        # 均值
        average_list = []
        average_index = 42
        average_bit_start = 12 + average_index*3
        # 方差
        variance_list = []
        variance_index = 44
        variance_bit_start = 12 + variance_index*3
        # ppm值
        ppm_list = []
        ppm_index = 5
        ppm_bit_start = 12 + ppm_index*3
        # 校验差值A，旧串口协议，s值 - 报警阈值(仅未识别烟雾类型前有效)
        s_value_index_a = 9
        s_value_a_bit_start = 12 + s_value_index_a * 3
        alarm_level_index_a = 3
        alarm_level_a_bit_start = 12 + alarm_level_index_a * 3
        # 校验差值B，旧串口协议，s值 - 报警阈值(仅未识别烟雾类型前有效)
        s_value_index_b = 26
        s_value_b_bit_start = 12 + s_value_index_b * 3
        alarm_level_index_b = 20
        alarm_level_b_bit_start = 12 + alarm_level_index_b * 3
        # 混合烟数组
        mix_smoke_value_list = []

        start_calc_flag = False
        mix_test_start_calc_flag = False
        mix_test_end_mark_flag = False
        alarm_flag = False
        start_mark_flag = False

        last_mix_test_dev_name_num = None
        last_dev_name_num = None
        dev_name = None
        dev_column_min = None
        dev_column_max = None
        dev_column_mean = None
        test_name = None

        for log_file in log_files:
            file_full_path = str(path + '/' + log_file)
            with open(file_full_path, "r", encoding="utf-8") as file:
                for line in file.readlines():
                    # print(line.strip())  # 使用strip()方法去除换行符
                    sub_string_start = "Graph: mark start "
                    sub_string_end = "Graph: mark end "
                    if sub_string_start in line: #开始标记，获取设备编号
                        smoke_box_keyword = ['标准烟雾', 'UL烟雾', '灵敏度', '烟箱']
                        pu_test_keyword = ['PU', '聚氨酯', '海绵', '聚']
                        oil_test_keyword = ['烹饪', '油烟', '油']
                        mix_test_keyword = ['混', '混合', '混合烟']

                        if any(k in line for k in smoke_box_keyword):
                            test_name = "UL烟雾"
                        elif any(k in line for k in pu_test_keyword):
                            test_name = "PU烟雾"
                        elif any(k in line for k in oil_test_keyword):
                            test_name = "油烟"
                        elif any(k in line for k in mix_test_keyword):
                            test_name = "混合烟"
                        else:
                            test_name = "Null"

                        start_time = self.get_timestamp(str(line))

                        pattern1 = r'\d+\#'
                        dev_name = re.search(pattern1, str(line))  # 使用 search 更稳妥，或者确保 findall 有结果
                        print('start1:', dev_name)
                        if dev_name:
                            start_mark_flag = True
                            dev_name_num = str(dev_name.group(0))
                            dev_column_min = dev_name_num + '_min'
                            dev_column_max = dev_name_num + '_max'
                            dev_column_mean = dev_name_num + '_mean'
                            if dev_column_min in pd_data.columns:
                                print('已经有：', dev_column_min)
                            else:
                                pd_data.loc[:, dev_column_min] = 0
                            if dev_column_max in pd_data.columns:
                                print('已经有：', dev_column_max)
                            else:
                                pd_data.loc[:, dev_column_max] = 0
                            if dev_column_mean in pd_data.columns:
                                print('已经有：', dev_column_mean)
                            else:
                                pd_data.loc[:, dev_column_mean] = 0
                            print("开始标记：", dev_name, f"({test_name})")
                        else:
                            sub_line = line.split(sub_string_start, maxsplit=1)
                            if sub_line[1] in line:
                                err_dev_name = sub_line[1]
                            self.text_edit.append("设备编号不符号要求:" + self.warning.format(f'{err_dev_name}⚠️'))
                            dev_name = None
                            continue
                    elif sub_string_end in line: #结束标记
                        end_time = self.get_timestamp(str(line))
                        if end_time - start_time > 60:
                            if test_name == "油烟":
                                test_type = '油烟-结束时红光增量'
                                pd_data.loc[test_type, dev_column_min] = increment_a_value
                                pd_data.loc[test_type, dev_column_max] = increment_a_value
                                pd_data.loc[test_type, dev_column_mean] = increment_a_value
                                print('油烟结束时红光增量=', increment_a_value)
                            elif test_name == "混合烟":
                                last_mix_test_dev_name_num = dev_name_num
                        elif end_time - start_time < 10:
                            if test_name == "混合烟":
                                if last_mix_test_dev_name_num == dev_name_num:
                                    last_mix_test_dev_name_num = None
                                    mix_test_end_mark_flag = True
                                    print(f'{dev_name_num}:混合烟测试结束')
                                else:
                                    start_calc_flag = False
                                    mix_test_start_calc_flag = False
                                    mix_test_end_mark_flag = False
                                    alarm_flag = False
                                    start_mark_flag = False

                                    last_mix_test_dev_name_num = None
                                    last_dev_name_num = None
                                    dev_name = None

                        print(f"{dev_name_num}结束标记")
                    elif dev_name is not None:
                        if "Receive" not in line:
                            continue
                        sub_line = line.split("Receive: ", maxsplit=1)
                        sub_line = sub_line[1].replace('"','')
                        if len(sub_line) > 30:
                            if (alarm_flag == False) and (int('0x' + sub_line[states_bit_start:states_bit_start+2], 16)
                                                                          == 0x07):
                                alarm_flag = True
                                print("报警了")

                            #增量比值
                            increment_ration_value = int('0x' + sub_line[increment_bit_start:increment_bit_start+2] +
                                                        sub_line[increment_bit_start + 3:increment_bit_start + 5], 16)
                            #增量A
                            increment_a_value = int('0x' + sub_line[increment_a_bit_start:increment_a_bit_start+2] +
                                                    sub_line[increment_a_bit_start + 3:increment_a_bit_start + 5], 16)
                            #增量B
                            increment_b_value = int('0x' + sub_line[increment_b_bit_start:increment_b_bit_start+2] +
                                                    sub_line[increment_b_bit_start + 3:increment_b_bit_start + 5], 16)
                            #比值变化率，有符号
                            byte_data = bytes.fromhex(sub_line[ration_bit_start:ration_bit_start+2] +
                                                            sub_line[ration_bit_start + 3:ration_bit_start + 5])
                            ration_of_change_value = int.from_bytes(byte_data, byteorder='big', signed=True)
                            #计数
                            rise_cnt_value = int('0x' + sub_line[rise_bit_start:rise_bit_start+2] +
                                                            sub_line[rise_bit_start + 3:rise_bit_start + 5], 16)
                            #均值，有符号
                            byte_data = bytes.fromhex(sub_line[average_bit_start:average_bit_start+2] +
                                                            sub_line[average_bit_start + 3:average_bit_start + 5])
                            average_value = int.from_bytes(byte_data, byteorder='big', signed=True)

                            #方差
                            variance_value = int('0x' + sub_line[variance_bit_start:variance_bit_start+2] +
                                                        sub_line[variance_bit_start + 3:variance_bit_start + 5], 16)
                            #ppm值
                            ppm_value = int('0x' + sub_line[ppm_bit_start:ppm_bit_start+2] +
                                                            sub_line[ppm_bit_start + 3:ppm_bit_start + 5], 16)

                            #如果增量比值=0，且校验开启标准=True，则结束计算（混合烟不满足increment_ration_value == 0）
                            if increment_ration_value == 0 and start_calc_flag == True:
                                start_calc_flag = False
                                if len(increment_ration_list) > 2:
                                    print("---------------------------------------")
                                    print(dev_name)
                                    dev_name = None
                                    if len(increment_ration_list) !=0:
                                        increment_ration_min = np.min(increment_ration_list)
                                        increment_ration_max = np.max(increment_ration_list)
                                        increment_ration_mean = int(np.mean(increment_ration_list))
                                        print("增量比值: 均值", increment_ration_mean,
                                              "最大值:", increment_ration_max, "最小值:", increment_ration_min)
                                        test_type = str(test_name) + '-增量比值'
                                        pd_data.loc[test_type, dev_column_min] = increment_ration_min
                                        pd_data.loc[test_type, dev_column_max] = increment_ration_max
                                        pd_data.loc[test_type, dev_column_mean] = increment_ration_mean

                                    if len(ration_of_change_list) != 0:
                                        ration_of_change_min = np.min(ration_of_change_list)
                                        ration_of_change_max = np.max(ration_of_change_list)
                                        ration_of_change_mean = int(np.mean(ration_of_change_list))
                                        print("比值变化率: 均值", ration_of_change_mean,
                                              "最大值:", ration_of_change_max, "最小值:", ration_of_change_min)
                                        test_type = str(test_name) + '-比值变化率'
                                        pd_data.loc[test_type, dev_column_min] = ration_of_change_min
                                        pd_data.loc[test_type, dev_column_max] = ration_of_change_max
                                        pd_data.loc[test_type, dev_column_mean] = ration_of_change_mean

                                    if len(rise_cnt_list) != 0:
                                        rise_cnt_min = np.min(rise_cnt_list)
                                        rise_cnt_max = np.max(rise_cnt_list)
                                        rise_cnt_mean = int(np.mean(rise_cnt_list))
                                        print("计数: 均值", rise_cnt_mean, "最大值:",
                                              rise_cnt_max, "最小值:", rise_cnt_min)
                                        test_type = str(test_name) + '-计数'
                                        pd_data.loc[test_type, dev_column_min] = rise_cnt_min
                                        pd_data.loc[test_type, dev_column_max] = rise_cnt_max
                                        pd_data.loc[test_type, dev_column_mean] = rise_cnt_mean

                                    if len(average_list) != 0:
                                        average_min = np.min(average_list)
                                        average_max = np.max(average_list)
                                        average_mean = int(np.mean(average_list))
                                        print("平均值: 均值", average_mean,
                                              "最大值:", average_max, "最小值:", average_min)
                                        test_type = str(test_name) + '-平均值'
                                        pd_data.loc[test_type, dev_column_min] = average_min
                                        pd_data.loc[test_type, dev_column_max] = average_max
                                        pd_data.loc[test_type, dev_column_mean] = average_mean

                                    if len(variance_list) != 0:
                                        variance_min = np.min(variance_list)
                                        variance_max = np.max(variance_list)
                                        variance_mean = int(np.mean(variance_list))
                                        print("方差: 均值", variance_mean, "最大值:",
                                              variance_max, "最小值:", variance_min)
                                        test_type = str(test_name) + '-方差'
                                        pd_data.loc[test_type, dev_column_min] = variance_min
                                        pd_data.loc[test_type, dev_column_max] = variance_max
                                        pd_data.loc[test_type, dev_column_mean] = variance_mean

                                    #校验L-D(A)
                                    print("通用-A通道校机差值:", cali_value_a)
                                    test_type = '通用-A通道校机差值'
                                    pd_data.loc[test_type, dev_column_min] = cali_value_a
                                    pd_data.loc[test_type, dev_column_max] = cali_value_a
                                    pd_data.loc[test_type, dev_column_mean] = cali_value_a

                                    #校验L-D(B)
                                    print("通用-B通道校机差值:", cali_value_b)
                                    test_type = '通用-B通道校机差值'
                                    pd_data.loc[test_type, dev_column_min] = cali_value_b
                                    pd_data.loc[test_type, dev_column_max] = cali_value_b
                                    pd_data.loc[test_type, dev_column_mean] = cali_value_b
                                    print("---------------------------------------")

                                increment_ration_list.clear()
                                ration_of_change_list.clear()
                                rise_cnt_list.clear()
                                average_list.clear()
                                variance_list.clear()
                                ppm_list.clear()
                            # 如果b通道增量>200，或者，校验开启标准=True，一直解析数据
                            elif increment_ration_value > 0:
                                if not start_calc_flag:
                                    start_calc_flag = True
                                else:#重置后第二次开始计算
                                    print("value:", increment_ration_value, increment_a_value, increment_b_value,
                                          ration_of_change_value, rise_cnt_value, average_value, variance_value,
                                          ppm_value)
                                    if test_name == '混合烟':
                                        #A通道>300后数组值
                                        if increment_a_value > 300:                 #红光增量>300
                                            if mix_test_start_calc_flag:
                                                mix_smoke_value_list.append(increment_a_value)
                                                # print(mix_smoke_value_list)
                                                if len(mix_smoke_value_list) >= 64:
                                                    # 去掉两个最大、最小值
                                                    mix_value_list = self.pop_list_max_min(mix_smoke_value_list)
                                                    mix_value_list = self.pop_list_max_min(mix_value_list)
                                                    # 得到最大最小值索引号
                                                    max_index = np.argmax(mix_value_list)
                                                    min_index = np.argmin(mix_value_list)
                                                    #去掉尾部
                                                    mix_smoke_value_list.pop(0)
                                                    # 当最小值的索引号 > 大于最大值索引号时
                                                    if min_index > max_index:
                                                        min_in_mix_smoke_list = np.min(mix_value_list)
                                                        pd_data.loc[
                                                            '混合烟-数组最小值', dev_column_min] = min_in_mix_smoke_list
                                                        pd_data.loc[
                                                            '混合烟-数组最小值', dev_column_max] = min_in_mix_smoke_list
                                                        pd_data.loc[
                                                            '混合烟-数组最小值', dev_column_mean] = min_in_mix_smoke_list
                                                        max_in_mix_smoke_list = np.max(mix_value_list)
                                                        pd_data.loc[
                                                            '混合烟-数组最大值', dev_column_min] = max_in_mix_smoke_list
                                                        pd_data.loc[
                                                            '混合烟-数组最大值', dev_column_max] = max_in_mix_smoke_list
                                                        pd_data.loc[
                                                            '混合烟-数组最大值', dev_column_mean] = max_in_mix_smoke_list
                                                        mix_test_start_calc_flag = False
                                                        print('混合烟-数组最小值: ', min_in_mix_smoke_list)
                                                        print('混合烟-数组最大值: ', max_in_mix_smoke_list)
                                        else:
                                            mix_test_start_calc_flag = True
                                            mix_smoke_value_list.clear()
                                            
                                        #求增量比值和差比值变化率
                                        if mix_test_end_mark_flag:  # 混合烟2次结束标记
                                            dev_name = None
                                            mix_test_end_mark_flag = False
                                            if len(increment_ration_list) > 10:
                                                increment_ration_min = np.min(increment_ration_list)
                                                increment_ration_max = np.max(increment_ration_list)
                                                increment_ration_mean = int(np.mean(increment_ration_list))
                                                print("增量比值: 均值", increment_ration_mean,
                                                      "最大值:", increment_ration_max, "最小值:",
                                                      increment_ration_min)
                                                test_type = str(test_name) + '-增量比值'
                                                pd_data.loc[test_type, dev_column_min] = increment_ration_min
                                                pd_data.loc[test_type, dev_column_max] = increment_ration_max
                                                pd_data.loc[test_type, dev_column_mean] = increment_ration_mean

                                            if len(ration_of_change_list) > 10:
                                                ration_of_change_min = np.min(ration_of_change_list)
                                                ration_of_change_max = np.max(ration_of_change_list)
                                                ration_of_change_mean = int(np.mean(ration_of_change_list))
                                                print("比值变化率: 均值", ration_of_change_mean,
                                                      "最大值:", ration_of_change_max, "最小值:",
                                                      ration_of_change_min)
                                                test_type = str(test_name) + '-比值变化率'
                                                pd_data.loc[test_type, dev_column_min] = ration_of_change_min
                                                pd_data.loc[test_type, dev_column_max] = ration_of_change_max
                                                pd_data.loc[test_type, dev_column_mean] = ration_of_change_mean

                                            increment_ration_list.clear()
                                            ration_of_change_list.clear()
                                        else:
                                            increment_ration_list.append(increment_ration_value)
                                            ration_of_change_list.append(ration_of_change_value)
                                    else:#非混合烟
                                        increment_ration_list.append(increment_ration_value)
                                        ration_of_change_list.append(ration_of_change_value)

                                        if increment_a_value < 380:
                                            rise_cnt_list.append(rise_cnt_value)
                                            average_list.append(average_value)
                                            variance_list.append(variance_value)
                                            ppm_list.append(ppm_value)
                            elif start_mark_flag:
                                start_calc_flag = False
                                start_mark_flag = False
                                #B通道初始增量
                                print("init increment_b_value=", increment_b_value)
                                pd_data.loc['通用-B通道初始增量', dev_column_min] = increment_b_value
                                pd_data.loc['通用-B通道初始增量', dev_column_max] = increment_b_value
                                pd_data.loc['通用-B通道初始增量', dev_column_mean] = increment_b_value
                                # 校机L-D差值(A)
                                s_value_a = int('0x' + sub_line[s_value_a_bit_start:s_value_a_bit_start + 2] +
                                                sub_line[s_value_a_bit_start + 3:s_value_a_bit_start + 5], 16)
                                alarm_level_a = int(
                                    '0x' + sub_line[alarm_level_a_bit_start:alarm_level_a_bit_start + 2] +
                                    sub_line[alarm_level_a_bit_start + 3:alarm_level_a_bit_start + 5], 16)
                                cali_value_a = s_value_a - alarm_level_a

                                # 校机L-D差值(B)
                                s_value_b = int('0x' + sub_line[s_value_b_bit_start:s_value_b_bit_start + 2] +
                                                sub_line[s_value_b_bit_start + 3:s_value_b_bit_start + 5], 16)
                                alarm_level_b = int(
                                    '0x' + sub_line[alarm_level_b_bit_start:alarm_level_b_bit_start + 2] +
                                    sub_line[alarm_level_b_bit_start + 3:alarm_level_b_bit_start + 5], 16)
                                cali_value_b = s_value_b - alarm_level_b
                                print("cali_L-D(A)", cali_value_a, "cali_L-D(B)", cali_value_b)

        #检查条件是否满足
        print("数据1:", pd_data)
        self.check_conditions(pd_set_data, pd_data)

        self.show_save_dialog()

    @staticmethod
    def pop_list_max_min(list_arg):
        seq = list(list_arg)
        seq.pop(seq.index(max(seq)))
        seq.pop(seq.index(min(seq)))
        return seq

    def check_conditions(self, set_df, df):
        print('开启条件刷选')
        for r_index, (index, row) in enumerate(df.iterrows()):
            result = set_df.loc[set_df['名称'] == str(index), '条件']
            if len(result):
                result_data = set_df.loc[set_df['名称'] == str(index), '条件'].values[0]
                if result_data == '是':
                    min_value = set_df.loc[set_df['名称'] == str(index), '最小值'].values[0]
                    max_value = set_df.loc[set_df['名称'] == str(index), '最大值'].values[0]
                    print(index, f'[{min_value}~{max_value}]')

                    for c_index, value in enumerate(row, 1):
                        if c_index%3 == 0:
                            print(c_index, row.iloc[c_index-3], row.iloc[c_index-2])
                            c_index_min = c_index - 3
                            c_index_max = c_index - 2
                            c_index_mean = c_index - 1
                            row_min = row.iloc[c_index_min]
                            row_max = row.iloc[c_index_max]
                            row_mean = row.iloc[c_index_mean]
                            if int(row_min) == int(row_max):
                                if  int(min_value) < int(row_mean) < int(max_value):
                                    print("value OK")
                                else:
                                    str_value = str(row_min) + '(F)'
                                    col_name = df.columns[c_index_min]
                                    df[col_name] = df[col_name].astype('object')
                                    df.iloc[r_index, c_index_min] = str_value
                                    col_name = df.columns[c_index_max]
                                    df[col_name] = df[col_name].astype('object')
                                    df.iloc[r_index, c_index_max] = str_value
                                    col_name = df.columns[c_index_mean]
                                    df[col_name] = df[col_name].astype('object')
                                    df.iloc[r_index, c_index_mean] = str_value
                            else:
                                min_value_err_flag = False
                                if int(row_min) < int(min_value):
                                    min_value_err_flag = True
                                    str_value = str(row_min)+'(F)'
                                    col_name = df.columns[c_index_min]
                                    df[col_name] = df[col_name].astype('object')
                                    df.iloc[r_index, c_index_min] = str_value
                                    print(str_value, r_index, c_index_min)
                                if int(row_max) > int(max_value) or \
                                                        (min_value_err_flag == True and int(row_min) >= int(row_max)):
                                    str_value = str(row_max)+'(F)'
                                    col_name = df.columns[c_index_max]
                                    df[col_name] = df[col_name].astype('object')
                                    df.iloc[r_index, c_index_max] = str_value
                                    print(str_value, r_index, c_index_max)
                    #修改需要条件判断的index标签
                    df = df.rename(index={str(index):str(index)+':['+str(min_value)+'~'+str(max_value)+']'+'(T)'})
        self.df = df
        # for idx_row, row in df.iterrows():
        #     print(df.index[idx_row])
            #
            # if row['条件'] == '是':
            #     print(f"{row['名称']}: {row['最小值']}, {row['最大值']}")


    def show_save_dialog(self):
        #获取当前时间作为文件名
        timestamp = time.time()
        local_time = time.localtime(timestamp)
        formatted_time = time.strftime('%Y-%m-%d_%H-%M-%S', local_time)

        # 使用 os.path.normpath 规范化路径，避免非法字符
        default_dir = os.path.normpath(QtCore.QDir.currentPath() + '/output/')
        default_file = self.cb.currentText() +'_' + formatted_time + '.xlsx'
        full_path = os.path.join(default_dir, default_file)  # 使用 os.path.join 处理路径分隔符

        print("保存路径:", default_dir)

        # 确保目录存在，exist_ok=True 避免重复创建错误
        os.makedirs(default_dir, exist_ok=True)

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            '保存 Excel 文件',
            full_path,
            'Excel Files (*.xlsx)'
        )

        if file_path:
            print("确认保存:",file_path)
            self.save_file(file_path)
        else:
            print("取消保存")

    def save_file(self, save_file):
        try:
            print(self.df)
            # 初始化 Workbook
            wb = Workbook()
            wb_s = wb.active
            columns_list = self.df.columns.to_list()

            # --- 样式定义 ---
            # 居中：水平居中，垂直居中
            align_center = Alignment(horizontal='center', vertical='center')
            # 字体：宋体，11号，加粗
            font_header = Font(name='宋体', size=14, bold=True)
            # 边框（可选，通常表头加个边框更好看）
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            # 填充色
            fill_header = PatternFill(start_color='0C9C0C', end_color='0C9C0C', fill_type='solid')
            fill_child_header = PatternFill(start_color='DDD9C4', end_color='DDD9C4', fill_type='solid')
            fill_warning = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

            # 靠左对齐
            left_align = Alignment(horizontal='right')

            # 1. 需要先写入值
            rows = dataframe_to_rows(self.df, index=True, header=True)
            for r_idx, row in enumerate(rows, 1):  # 从第1行开始计数
                if r_idx == 1:
                    for c_idx in range(0, len(columns_list)):
                        if c_idx > 1:
                            print("c_idx", c_idx, len(columns_list))
                            pattern = r'\d+\#'
                            match = re.search(pattern, str(columns_list[c_idx - 2]))  # 使用 search 更稳妥，或者确保 findall 有结果
                            if match:
                                wb_s.cell(row=1, column=c_idx, value=match.group(0))
                                cell = wb_s.cell(row=1, column=c_idx)   #写入数据时，同时设置单元格样式
                                cell.alignment = align_center
                                cell.font = font_header
                                cell.border = thin_border
                                cell.fill = fill_header
                elif r_idx == 2:
                    for c_idx in range(0, len(columns_list), 3):#for in循环， 起始值、终止值和步长
                        print("c_idx", c_idx, len(columns_list))
                        cell = wb_s.cell(row=2, column=c_idx + 2, value='最小值')
                        cell.border = thin_border
                        cell.fill = fill_child_header
                        cell = wb_s.cell(row=2, column=c_idx + 3, value='最大值')
                        cell.border = thin_border
                        cell.fill = fill_child_header
                        cell = wb_s.cell(row=2, column=c_idx + 4, value='平均值')
                        cell.border = thin_border
                        cell.fill = fill_child_header
                else:
                    for c_idx, value in enumerate(row, 1):
                        # 直接通过单元格坐标写入，而不是 append
                        cell = wb_s.cell(row=r_idx, column=c_idx, value=value)
                        if '(F)' in str(cell.value):
                            # 假设 pd_data['col'] 包含 'name(F)'
                            #pd_data['col名'] = pd_data['col名'].str.replace('(F)', '', regex=False)
                            #wb_s.cell(row=r_idx, column=c_idx, value=str(value).replace('(F)', ''))
                            cell.value = str(cell.value).replace('(F)', '')
                            cell.fill = fill_warning
                            cell.alignment = left_align

                            header_cell = wb_s.cell(row=1, column=c_idx)
                            header_cell.fill = fill_warning
                        if c_idx > 1 and (c_idx-2)%3 == 0:
                            # cell = wb_s.cell(row=r_idx, column=c_idx)
                            cell.border = Border(left=Side(style='thin'), right=None, top=None, bottom=None)
                    # for c_idx in range(0, len(columns_list), 3):  # for in循环， 起始值、终止值和步长,2 5 8 11
                    #     cell = wb_s.cell(row=r_idx, column=c_idx + 2)
                    #     cell.border = Border(left=Side(style='thin'), right=None, top=None, bottom=None)

            # 2. 必须执行1写入后，才能合并单元格，写入设备名称
            columns_array = columns_list[::3]
            for index in range(len(columns_array)):
                    wb_s.merge_cells(None, 1, index*3+2, 1, index*3+4)

            # 3. 合并显示设备号
            wb_s.merge_cells('A1:A2')
            wb_s['A1'] = '设备号'
            cell = wb_s['A1']
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.fill = PatternFill(start_color='C5D9F1', end_color='C5D9F1', fill_type='solid')

            # 4. 遍历第一列A，将徐娅测试的index名称标记颜色
            for cell in wb_s['A'][1:]:
                if '(T)' in str(cell.value):
                    cell.fill = PatternFill(start_color='C5D9F1', end_color='C5D9F1', fill_type='solid')
                    cell.value = str(cell.value).replace('(T)', '')
                    print(cell.value)
                # print(cell.value)


            # 5. 设置首列A的宽度
            wb_s.column_dimensions['A'].width = 32

            wb.save(save_file)
            wb.close()
            print("✅ 文件写入成功！")
            print(self.df)
        except Exception as e:
            print(f"❌ 写入失败，错误信息: {e}")
            self.text_edit.append(self.error.format("写入excel失败，确保文件在关闭状态❌"))

        # 请求完成
        file_name = os.path.basename(save_file)
        print("文件名",file_name)  # 输出: file.txt
        self.text_edit.append(self.valid.format("刷选完成✅ "))
        self.text_edit.append(self.valid.format(f"输出文件:{file_name}✅ "))
        try:
            os.startfile(save_file)  # Windows系统可用
        except Exception as e:
            print(f"❌ 打开excel失败: {e}")
            self.text_edit.append(self.error.format("打开excel失败❌"))